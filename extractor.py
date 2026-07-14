import re
import sys
import openpyxl
import pdfplumber

# =====================================================================
# 1) CONFIGURAÇÃO - MAPEAMENTO DO SEU TEMPLATE DE EXCEL
# =====================================================================
EXCEL_MAP = {
    "mrc": {
        "sheet": None,
        "start_row": 6,
        "max_row": 21,
        "cols": {"product": "C", "sku": "D", "qty": "E", "price": "H"},
    },
    "usage": {
        "sheet": None,
        "start_row": 31,
        "max_row": 44,
        "cols": {"product": "C", "sku": "D", "qty": "E", "price": "H"},
    },
    "nrc": {
        "sheet": None,
        "start_row": 64,
        "max_row": 90,
        "cols": {"product": "C", "sku": "D", "qty": "E", "price": "H"},
    },
    "connectivity_mrc": {
        "sheet": None,
        "start_row": 95,
        "max_row": 98,
        "cols": {"product": "C", "sku": "D", "qty": "E", "price": "H"},
    },
    "connectivity_nrc": {
        "sheet": None,
        "start_row": 103,
        "max_row": 106,
        "cols": {"product": "C", "sku": "D", "qty": "E", "price": "H"},
    },
}

# Títulos de seção atualizados (Sempre em MAIÚSCULO para validação do motor)
TITLES = {
    # Originais
    "SOFTWARE MRC PRODUCTS": "mrc",
    "SOFTWARE NRC PRODUCTS": "nrc",
    "SOFTWARE USAGE PRODUCTS": "usage",
    "DEDICATED NETWORK CONNECTIVITY MRC": "connectivity_mrc",
    "DEDICATED NETWORK CONNECTIVITY NRC": "connectivity_nrc",
    
    # Novas Seções Adicionadas
    "PER USER/UNIT SUBSCRIPTIONS": "mrc",
    "PER BU SUBSCRIPTIONS": "nrc",
    "CONSULTING": "nrc",
    "USAGE PRODUCTS": "usage",
    "IMPLEMENTATION & TRAINING": "nrc",
    "MONTHLY LOOP QUOTE SUBSCRIPTIONS": "connectivity_mrc",
    "LOOP QUOTE SETUP & ACTIVATION": "connectivity_nrc",
}

LAYOUT = {
    "mrc": "software",
    "nrc": "software",
    "usage": "software",
    "connectivity_mrc": "connectivity",
    "connectivity_nrc": "connectivity",
}


# =====================================================================
# 2) TRATAMENTO DE TEXTO E CONVERSÕES
# =====================================================================
def norm(s):
    """Remove quebras de linha internas de célula e espaços duplicados."""
    return re.sub(r"\s+", " ", (s or "").replace("\n", " ")).strip()

def clean_sku(s):
    """Remove espaços internos do SKU."""
    if not s:
        return ""
    s = norm(s)
    s = re.sub(r"\s+", "", s)
    return s.upper()

def parse_price(s):
    """Converte strings monetárias para float."""
    if s is None:
        return None
    s = re.sub(r"\s+", "", s) # remove espaços e quebras de linha primeiro
    s = re.sub(r"(?i)(BRL|USD)", "", s)
    s = s.replace("R$", "").replace("US$", "").replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None
    
def parse_qty(s):
    """Converte strings de quantidade para int ou float."""
    if s is None:
        return None
    s = s.replace(",", "").strip()
    if s == "":
        return None
    try:
        v = float(s)
        return int(v) if v == int(v) else v
    except ValueError:
        return None


# =====================================================================
# 3) EXTRAÇÃO EXCLUSIVA DO PDF
# =====================================================================
def extract_pdf(pdf_path):
    data = {key: [] for key in set(TITLES.values())}
    current_section = None
    current_title = None
    finished_titles = set()   # <-- por título, não por seção

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row:
                        continue
                    first_cell = norm(row[0]).upper()

                    if first_cell in TITLES:
                        if first_cell in finished_titles:
                            current_section = None
                        else:
                            current_section = TITLES[first_cell]
                            current_title = first_cell
                        continue

                    if first_cell in ("PRODUCT CODE", "PRODUCT", "CATALOG ID"):
                        continue

                    if current_section is None:
                        continue

                    if all(c is None or norm(c) == "" for c in row):
                        continue

                    if len(row) < 4:
                        continue

                    qty = parse_qty(row[-3])
                    price = parse_price(row[-2])

                    if qty is None or price is None:
                        if current_title is not None:
                            finished_titles.add(current_title)
                        current_section = None
                        continue

                    layout = LAYOUT[current_section]
                    if layout == "software":
                        sku = clean_sku(row[0])
                        product = " ".join(norm(c) for c in row[1:-3] if norm(c))
                    else:
                        sku = ""
                        product = norm(row[0])

                    if not product:
                        finished_titles.add(current_title)
                        current_section = None
                        continue

                    data[current_section].append(
                        {"sku": sku, "product": product, "qty": qty, "price": price}
                    )

    return data


# =====================================================================
# 4) ESCRITA NO TEMPLATE EXCEL
# =====================================================================
def write_to_excel(data, template_path, output_path):
    wb = openpyxl.load_workbook(template_path)

    for section, items in data.items():
        cfg = EXCEL_MAP[section]
        ws = wb[cfg["sheet"]] if cfg["sheet"] else wb.active
        cols = cfg["cols"]

        overflow = []
        for i, item in enumerate(items):
            row = cfg["start_row"] + i
            if row > cfg["max_row"]:
                overflow.append(item)
                continue
            ws[f"{cols['product']}{row}"] = item["product"]
            ws[f"{cols['sku']}{row}"] = item["sku"]
            ws[f"{cols['qty']}{row}"] = item["qty"]
            ws[f"{cols['price']}{row}"] = item["price"]

        if overflow:
            print(f"[AVISO] Seção '{section}' estourou o limite operacional estabelecido.")

    wb.save(output_path)


def main():
    if len(sys.argv) != 4:
        print("Uso: python extractor.py <cotacao.pdf> <template.xlsx> <saida.xlsx>")
        sys.exit(1)

    pdf_path, template_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]
    print(f"Processando arquivo: {pdf_path}")
    data = extract_pdf(pdf_path)
    write_to_excel(data, template_path, output_path)
    print("Sucesso.")


if __name__ == "__main__":
    main()